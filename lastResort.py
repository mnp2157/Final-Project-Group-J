import mysql.connector
from openpyxl import load_workbook

# Load the Excel workbook 'LastResortHotel' 
hotel_workbook = 'LastResortHotel.xlsx'
data_worksheet = load_workbook(filename=hotel_workbook, data_only=True)

# function to get all the cell values
def get_cell_value(sheet, row, col):
    return sheet.cell(row=row, column=col).value

# Establish a connection to the MySQL database
try:
    mydb = mysql.connector.connect(
        host="localhost",
        user="root",
        password="@Chococat15!",
        database="last_resort"  
    )
    cursor = mydb.cursor()

    # Insert data_worksheet for each table
    table_sheet_mapping = {
        'sleeping_room': 'sleeping_room',
        'staff': 'staff',
        'customer': 'customer',
        'sleeping_room_reservation': 'sleeping_room_reservation',
        'sleeping_room_guest': 'sleeping_room_guest',
        'access_card': 'access_card',
        'meeting_room': 'meeting_room',
        'meeting_room_reservation': 'meeting_room_reservation',
        'meeting_room_guest_list': 'meeting_room_guest_list',
        'event': 'event',
        'event_list': 'event_list',
        'service_usage': 'service_usage',
        'sleeping_room_cost': 'sleeping_room_cost',
        'meeting_room_cost': 'meeting_room_cost',
        'receipt': 'receipt',
        'advance_deposit': 'advance_deposit',
    }

    for sheet_name, table_name in table_sheet_mapping.items():
        if sheet_name in data_worksheet.sheetnames:
            sheet = data_worksheet[sheet_name]

            # Filter out None values from the list of columns
            columns = [sheet.cell(row=1, column=col).value for col in range(1, sheet.max_column + 1) if
                       sheet.cell(row=1, column=col).value is not None]

            placeholders = ', '.join(['%s'] * len(columns))
            insert_query = f"INSERT INTO {table_name} ({', '.join(columns)}) VALUES ({placeholders})"

            data_row = 2  # because the data_worksheet starts at row 2
            while get_cell_value(sheet, data_row, 1) is not None:
                values = [get_cell_value(sheet, data_row, col) for col in range(1, sheet.max_column + 1) if
                          sheet.cell(row=1, column=col).value is not None]
                cursor.execute(insert_query, values)
                mydb.commit()
                data_row += 1

            print(f"Data inserted into {table_name} successfully.")
        else:
            print(f"Sheet {sheet_name} not found in the Excel file.")

except mysql.connector.Error as err:
    print(f"Error: {err}")
finally:
    if cursor:
        cursor.close()
    if mydb:
        mydb.close()
